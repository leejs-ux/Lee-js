import streamlit as st
import pandas as pd
import ezdxf
import os
import tempfile
from io import BytesIO
import openpyxl
import json
import re
import time
import google.generativeai as genai
import gspread

st.set_page_config(page_title="2D DXF 자동 견적 시스템", page_icon="⚙️", layout="wide")

# =========================================================================
# 💡 구글 시트(DB) 연결 로직
# =========================================================================
SHEET_NAME = "견적프로그램_DB"

@st.cache_resource
def init_gspread():
    if "google_credentials" in st.secrets:
        try:
            cred_dict = json.loads(st.secrets["google_credentials"], strict=False)
            gc = gspread.service_account_from_dict(cred_dict)
            return gc
        except Exception as e:
            st.error(f"⚠️ 구글 시트 인증 에러 (JSON 출입증을 확인하세요): {e}")
            return None
    else:
        st.error("⚠️ 스트림릿 금고(Secrets)에 google_credentials 암호가 없습니다.")
        return None

gc = init_gspread()

# =========================================================================
# 🔑 API 키 자동 저장/불러오기 로직
# =========================================================================
API_KEY_FILE = "api_key.txt"
def load_api_key():
    if os.path.exists(API_KEY_FILE):
        with open(API_KEY_FILE, "r") as f:
            return f.read().strip()
    return ""

def save_api_key(key):
    with open(API_KEY_FILE, "w") as f:
        f.write(key.strip())

st.sidebar.header("🔑 AI 설정")
saved_key = load_api_key()
api_key = st.sidebar.text_input("Gemini API Key를 입력하세요", value=saved_key, type="password")

if st.sidebar.button("💾 API 키 저장하기"):
    if api_key:
        save_api_key(api_key)
        st.sidebar.success("✅ 키가 컴퓨터에 저장되었습니다!")
    else:
        st.sidebar.error("키를 입력한 후 저장해주세요.")

st.title("⚙️ 2D DXF 기반 하이브리드 자동/반자동 견적 시스템")
st.markdown("---")

def safe_float(value):
    try:
        if isinstance(value, (int, float)):
            return float(value)
        num_str = re.sub(r'[^0-9.]', '', str(value))
        return float(num_str) if num_str else 0.0
    except:
        return 0.0

# =========================================================================
# 1. 기준 단가표 관리 (구글 시트 연동)
# =========================================================================
# 기본값 (시트가 비어있거나 에러날 때 사용)
default_material_db = pd.DataFrame({
    '재질': ['SS400', 'S45C', 'SPCC(레이져)', 'SM45C', 'AL2017', 'AL5052고베판', 'AL6061판재', 'AL7075', 'SUS304', 'BS(신주)', 'MC 나이론', '아세탈', '테프론', 'PC (국산)', 'PUR.'],
    'KG당 단가': [2400, 2400, 1500, 2400, 25000, 9300, 8000, 11500, 7650, 10000, 12000, 15000, 40000, 10000, 15000],
    '비중': [8.0, 8.0, 8.0, 8.0, 2.8, 2.8, 2.8, 2.8, 8.0, 8.0, 1.6, 1.41, 2.2, 1.2, 1.5]
})

default_post_db = pd.DataFrame({
    '표면처리': ['W-Anodizing', 'B-Anodizing', 'H-Anodizing', 'SOFT ANODIZING', '무전해니켈(ST)', '무전해니켈(AL)', '크롬도금', '전해연마', '아연도금', '흑색경질', 'POLISHING', 'PAINT'],
    'KG당 단가': [1500, 2500, 6000, 3000, 2500, 6000, 1500, 1500, 800, 7000, 2500, 600]
})

if 'material_db' not in st.session_state:
    if gc:
        try:
            sh = gc.open(SHEET_NAME)
            ws_m = sh.worksheet("material_db")
            data_m = ws_m.get_all_records()
            if data_m:
                st.session_state.material_db = pd.DataFrame(data_m)
            else:
                ws_m.update([default_material_db.columns.values.tolist()] + default_material_db.astype(str).values.tolist())
                st.session_state.material_db = default_material_db
        except Exception:
            st.session_state.material_db = default_material_db
    else:
        st.session_state.material_db = default_material_db

if 'post_db' not in st.session_state:
    if gc:
        try:
            sh = gc.open(SHEET_NAME)
            ws_p = sh.worksheet("post_db")
            data_p = ws_p.get_all_records()
            if data_p:
                st.session_state.post_db = pd.DataFrame(data_p)
            else:
                ws_p.update([default_post_db.columns.values.tolist()] + default_post_db.astype(str).values.tolist())
                st.session_state.post_db = default_post_db
        except Exception:
            st.session_state.post_db = default_post_db
    else:
        st.session_state.post_db = default_post_db

if 'parsed_df' not in st.session_state:
    st.session_state.parsed_df = pd.DataFrame()
if 'uploaded_file_names' not in st.session_state:
    st.session_state.uploaded_file_names = []

with st.expander("📊 1. 기준 단가표 관리 (클릭하여 펼치기)"):
    col1, col2 = st.columns(2)
    with col1:
        edited_material = st.data_editor(st.session_state.material_db, num_rows="dynamic", use_container_width=True)
    with col2:
        edited_post = st.data_editor(st.session_state.post_db, num_rows="dynamic", use_container_width=True)
    
    if st.button("💾 변경된 단가표 영구 저장하기 (구글 시트 연동)"):
        st.session_state.material_db = edited_material
        st.session_state.post_db = edited_post
        if gc:
            try:
                sh = gc.open(SHEET_NAME)
                ws_m = sh.worksheet("material_db")
                ws_m.clear()
                ws_m.update([edited_material.columns.values.tolist()] + edited_material.astype(str).values.tolist())
                
                ws_p = sh.worksheet("post_db")
                ws_p.clear()
                ws_p.update([edited_post.columns.values.tolist()] + edited_post.astype(str).values.tolist())
                
                st.success("✅ 구글 스프레드시트에 단가표가 완벽하게 동기화(저장)되었습니다!")
            except Exception as e:
                st.error(f"⚠️ 구글 시트 저장 실패: {e}")
        else:
            st.warning("⚠️ 구글 시트와 연결되지 않아 임시 저장만 되었습니다.")

st.markdown("---")

# =========================================================================
# 🤖 진짜 AI (Gemini) 파싱 함수 - 범용성 대폭 강화!
# =========================================================================
def analyze_with_gemini(filename, text_data, geometry_info, api_key):
    genai.configure(api_key=api_key)
    
    target_model_name = ""
    try:
        available_models = genai.list_models()
        for m in available_models:
            if 'generateContent' in m.supported_generation_methods:
                if 'gemini' in m.name.lower():
                    target_model_name = m.name
                    if 'flash' in m.name.lower():
                        break
    except Exception as e:
        st.error(f"⚠️ [{filename}] AI 모델 탐색 에러: {e}")
        return {"도면번호": filename, "품명": "분석 실패", "재질": "미정", "수량": 1, "가로": 0, "세로": 0, "두께": 0, "후처리": "없음", "비고": f"AI 탐색 에러: {e}"}
    
    if not target_model_name:
        target_model_name = "gemini-1.5-flash"
        
    model = genai.GenerativeModel(target_model_name)
    
    # 💡 도면 폼이 바뀌어도 확실하게 데이터를 뽑아내도록 명령어를 대폭 강화했습니다.
    prompt = f"""
    당신은 대한민국 가공업계 베테랑 도면 분석 전문가입니다. DXF 도면에서 추출한 '무작위 텍스트 덩어리'를 정밀 해독하여 견적 핵심 데이터를 JSON 형식으로 추출해야 합니다.

    **🚨 도면 폼(양식)이 다르더라도 문제없이 데이터를 찾기 위한 엄격 지침:**

    1.  **표제란(Title Block) 및 부품표(Parts List) 우선 탐색:** 도면 오른쪽 하단이나 상단에 위치한 '표(Table)' 형태의 텍스트 밀집 구역을 먼저 읽으십시오. 용어(헤더) 아래에 있는 값을 찾으세요.
    2.  **용어 범용성(Synonyms) 고려:**
        * **재질:** '재질', 'MAT'L', 'MATERIAL', 'MATE', 'MAT' 등을 찾고 그 근처 값을 읽으십시오. (예: "MC (흑색)" -> 'MC' 또는 'MC 나이론'으로 매핑)
        * **규격/치수:** '규격', '규 격', 'SPEC.', 'SIZE', 'DIMENSION' 등을 찾으십시오. 특히 키워드 근처에서 **"숫자x숫자x숫자"** 또는 "숫자*숫자*숫자" 패턴(예: 35X130X360)을 발견하면 순서대로 가로, 세로, 두께(또는 임의)로 간주하고 숫자만 추출하십시오. 이것은 치수선 값보다 우선합니다.
        * **수량:** '수량', 'Q'TY', 'QTY', 'QUANTITY' 등을 찾고 그 아래 값을 읽으십시오. 파일 갯수가 아니라 도면에 적힌 수량을 찾아야 합니다.
    3.  **데이터 정리 및 매핑:**
        * 수량, 가로, 세로, 두께는 오직 **'숫자'**만 적으십시오. (문자 제외)
        * 재질은 가장 적합한 표준 명칭(예: MC, SUS304, AL6061)을 추론하여 적으십시오.

    [제공된 데이터]
    -   파일명: '{filename}'
    -   도면 추출 기하 정보: {geometry_info}
    -   도면 추출 텍스트: {text_data}

    이 정보를 종합하여 **아래 JSON 형식으로만** 완벽하게 대답하십시오.

    {{
        "도면번호": "문자열 (DWG.NO. 칸 등 탐색)",
        "품명": "문자열 (TITLE 칸 등 탐색, 예: V-Block)",
        "재질": "문자열 (MAT'L 칸 탐색, 없으면 '미정')",
        "수량": 정수 (Q'TY 칸 탐색, 찾을 수 없으면 1),
        "가로": 숫자 (SPEC 칸의 'AxBxC' 패턴 또는 최대치수, 찾을 수 없으면 0),
        "세로": 숫자 (SPEC 칸의 'AxBxC' 패턴 또는 두번째치수, 찾을 수 없으면 0),
        "두께": 숫자 (SPEC 칸의 'AxBxC' 패턴 또는 두께 치수, 찾을 수 없으면 0),
        "후처리": "문자열 (주서란 등 탐색, 없으면 '없음')",
        "비고": "도면의 특징(예: Bending, Tap 정보), 추출 기하 정보 요약 및 주서 요약"
    }}
    """
    
    try:
        response = model.generate_content(prompt)
        result_text = response.text.strip()
        
        if result_text.startswith("```json"):
            result_text = result_text[7:-3].strip()
        elif result_text.startswith("```"):
            result_text = result_text[3:-3].strip()
            
        parsed_data = json.loads(result_text)
        return parsed_data
    except Exception as e:
        st.error(f"⚠️ [{filename}] AI 분석 중 에러 발생: {e}")
        return {"도면번호": filename, "품명": "분석 실패", "재질": "미정", "수량": 1, "가로": 0, "세로": 0, "두께": 0, "후처리": "없음", "비고": f"생성 에러: {e}"}

# =========================================================================
# 2. DXF 업로드 및 실시간 AI 분석 로직
# =========================================================================
st.subheader("2. DXF 도면 업로드 및 AI 분석")
uploaded_files = st.file_uploader("📂 DXF 도면들을 드래그 앤 드롭 하세요.", type=['dxf'], accept_multiple_files=True)

if uploaded_files:
    if not api_key:
        st.warning("👈 왼쪽 사이드바에 Gemini API Key를 먼저 입력해 주세요!")
    else:
        current_file_names = [f.name for f in uploaded_files]
        
        if st.session_state.uploaded_file_names != current_file_names:
            parsed_results = []
            with st.spinner("🤖 베테랑 AI가 새로운 명령어 지침에 따라 도면 폼을 분석하고 데이터를 추출하고 있습니다..."):
                for idx, file in enumerate(uploaded_files):
                    with tempfile.NamedTemporaryFile(delete=False, suffix=".dxf") as tmp_file:
                        tmp_file.write(file.getvalue())
                        tmp_path = tmp_file.name
                    
                    try:
                        doc = ezdxf.readfile(tmp_path)
                        msp = doc.modelspace()
                        
                        extracted_texts = [e.dxf.text for e in msp.query('TEXT MTEXT') if hasattr(e.dxf, 'text') and e.dxf.text]
                        # AI에게 더 많은 문맥을 제공하기 위해 띄어쓰기를 더 명확히 유지
                        clean_texts = " | ".join([t.strip() for t in extracted_texts if t.strip()])
                        
                        tolerance_keywords = ['±', '%%p', '+', '-', 'H7', 'h7']
                        num_tols = sum(1 for t in extracted_texts if any(k in t for k in tolerance_keywords))
                        
                        num_holes = len(msp.query('CIRCLE'))
                        num_dims = len(msp.query('DIMENSION'))
                        geometry_info = f"원(구멍) 갯수: {num_holes}개, 치수기입 갯수: {num_dims}개, 공차추정: {num_tols}건"
                        
                        ai_result = analyze_with_gemini(file.name, clean_texts, geometry_info, api_key)
                        
                        # AI가 가져온 숫자를 safe_float로 세탁
                        w = safe_float(ai_result.get("가로", 0))
                        h = safe_float(ai_result.get("세로", 0))
                        t = safe_float(ai_result.get("두께", 0))
                        qty = ai_result.get("수량", 1)
                        if not isinstance(qty, int): # 수량이 정수가 아니면 1로 강제
                             qty = 1
                        
                        ai_result["가로"] = w
                        ai_result["세로"] = h
                        ai_result["두께"] = t
                        ai_result["수량"] = qty
                        
                        mat_name = ai_result.get("재질", "미정")
                        post_name = ai_result.get("후처리", "없음")
                        
                        # 💡 재질 단가표 매핑 개선 (MC 흑색 같은 값을 DB의 'MC 나이론'에 매핑 시도)
                        mat_info = pd.DataFrame()
                        for idx, db_row in st.session_state.material_db.iterrows():
                             if str(db_row['재질']).lower() in str(mat_name).lower(): # 예: 'mc' 가 'mc 나이론'에 포함되면
                                  mat_info = st.session_state.material_db.iloc[[idx]]
                                  break
                        
                        if mat_info.empty: # 못 찾았으면 원본 이름으로 시도
                             mat_info = st.session_state.material_db[st.session_state.material_db['재질'] == mat_name]

                        if not mat_info.empty:
                            weight_ratio = mat_info['비중'].values[0]
                            mat_price_per_kg = mat_info['KG당 단가'].values[0]
                            volume = w * h * t 
                            # 비중 계산 (mm단위 치수이므로 10^6으로 나눔)
                            weight = volume * weight_ratio / 1000000 
                            ai_result["소재비"] = int(weight * mat_price_per_kg)
                        else:
                            ai_result["소재비"] = 0
                        
                        post_info = st.session_state.post_db[st.session_state.post_db['표면처리'] == post_name]
                        if not post_info.empty:
                            post_price_per_kg = post_info['KG당 단가'].values[0]
                            ai_result["후처리비"] = int(weight * post_price_per_kg) if 'weight' in locals() else 0
                        else:
                            ai_result["후처리비"] = 0
                        
                        ai_result["가공비(수동입력)"] = 0 
                        # 총합계 계산 (수량 곱하기는 엑셀 발행 시 진행)
                        ai_result["최종합계"] = ai_result["소재비"] + ai_result["후처리비"]
                        
                        parsed_results.append(ai_result)
                    
                    except Exception as e:
                        st.error(f"{file.name} 처리 중 오류: {e}")
                    finally:
                        os.remove(tmp_path)
                    
                    if idx < len(uploaded_files) - 1:
                        time.sleep(3) 
            
            st.session_state.parsed_df = pd.DataFrame(parsed_results)
            st.session_state.uploaded_file_names = current_file_names

        st.success("✅ 새로운 베테랑 지침에 따른 AI 도면 분석 및 기초 단가 계산이 완료되었습니다!")

        # 💡 과거 이력 조회 (구글 시트 연동)
        if gc and not st.session_state.parsed_df.empty:
            try:
                sh = gc.open(SHEET_NAME)
                ws_q = sh.worksheet("Quote_Database")
                records = ws_q.get_all_records()
                if records:
                    history_db = pd.DataFrame(records)
                    for idx, row in st.session_state.parsed_df.iterrows():
                        drw_no = str(row.get('도면번호', ''))
                        if not drw_no: continue
                        history_db['도면번호_str'] = history_db['도면번호'].astype(str)
                        matches = history_db[history_db['도면번호_str'] == drw_no]
                        if not matches.empty:
                            last_quote = matches.iloc[-1]
                            st.warning(f"🕒 **과거 이력 발견!** [{drw_no}] 도면은 구글 시트에 기존 견적 이력이 있습니다. \n"
                                       f"👉 **이전 기록:** 가공비 {last_quote.get('가공비(수동입력)', 0):,}원 / 최종합계 {last_quote.get('최종합계', 0):,}원 (비고: {last_quote.get('비고', '없음')})")
            except Exception:
                pass 

        if not st.session_state.parsed_df.empty:
            st.markdown("---")
            st.subheader("3. 📝 최종 견적 검토 및 데이터 수정")
            st.info("💡 모든 칸을 자유롭게 수정하실 수 있습니다.")
            
            edited_df = st.data_editor(
                st.session_state.parsed_df,
                disabled=["최종합계"], 
                hide_index=True,
                use_container_width=True,
                key="quote_editor" 
            )
            
            final_df = edited_df.copy()
            final_df["최종합계"] = final_df["소재비"] + final_df["후처리비"] + final_df["가공비(수동입력)"]

            st.markdown(f"### 💰 전체 프로젝트 총 견적액 (수량 미반영 개당 단가): **{final_df['최종합계'].sum():,} 원**")

            st.markdown("---")
            st.subheader("4. 💾 견적 확정 및 엑셀 다운로드")
            
            if st.button("🚀 견적 확정 및 엑셀 폼 발행하기"):
                if gc:
                    try:
                        sh = gc.open(SHEET_NAME)
                        ws_q = sh.worksheet("Quote_Database")
                        data_q = ws_q.get_all_values()
                        if not data_q: 
                            ws_q.update([final_df.columns.values.tolist()] + final_df.astype(str).values.tolist())
                        else: 
                            ws_q.append_rows(final_df.astype(str).values.tolist())
                        st.success(f"✅ DB 누적 완료!")
                    except Exception as e:
                        st.error(f"⚠️ 구글 시트 누적 저장 실패: {e}")
                
                # 엑셀 다운로드 (최종 양식에는 수량을 곱해서 발행)
                template_path = "견적서.xlsx"
                try:
                    wb = openpyxl.load_workbook(template_path)
                    ws = wb["견적서(을지)"] if "견적서(을지)" in wb.sheetnames else wb.active
                    
                    start_row = 7
                    for index, row in final_df.iterrows():
                        current_row = start_row + index
                        size_spec = f"{row['가로']} x {row['세로']} x {row['두께']}"
                        qty = row['수량']
                        
                        ws.cell(row=current_row, column=1).value = index + 1
                        ws.cell(row=current_row, column=2).value = row['도면번호']
                        ws.cell(row=current_row, column=3).value = row['품명']
                        ws.cell(row=current_row, column=4).value = size_spec
                        ws.cell(row=current_row, column=6).value = row['후처리']
                        ws.cell(row=current_row, column=7).value = qty
                        # 최종 발행 시 개당 단가에 수량을 곱해서 엑셀에 기입
                        ws.cell(row=current_row, column=8).value = row['소재비'] * qty
                        ws.cell(row=current_row, column=9).value = row['가공비(수동입력)'] * qty
                        ws.cell(row=current_row, column=10).value = row['후처리비'] * qty
                        ws.cell(row=current_row, column=16).value = row['비고']

                    output = BytesIO()
                    wb.save(output)
                    excel_data = output.getvalue()
                    
                    st.download_button(
                        label="📊 회사 양식 최종 엑셀 다운로드 (.xlsx)",
                        data=excel_data,
                        file_name="최종견적서_발행.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
                    )
                except Exception as e:
                    st.error(f"⚠️ 엑셀 템플릿 처리 중 오류: {e}")
